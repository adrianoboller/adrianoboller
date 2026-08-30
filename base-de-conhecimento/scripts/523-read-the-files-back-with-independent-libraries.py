# Read the files back with independent libraries
# 28/08 16:57

import openpyxl, datetime, docx
w = openpyxl.load_workbook('$SC/clientes.xlsx')
s = w.active
print("XLSX ------------------------------------------------")
print(" aba:", s.title, "| autofiltro:", s.auto_filter.ref, "| painel congelado:", s.freeze_panes)
for r in s.iter_rows(min_row=1, max_row=8):
    print("  ", [(c.value if not isinstance(c.value,(datetime.date,datetime.datetime)) else c.value.isoformat()) for c in r])
d = s.cell(row=5, column=5)
print(" nascimento e mesmo data?", type(d.value).__name__, "->", d.value, "| formato:", d.number_format)
lim = s.cell(row=5, column=4)
print(" limite e mesmo numero? ", type(lim.value).__name__, "->", lim.value, "| formato:", lim.number_format)
cab = s.cell(row=4, column=1)
print(" cabecalho pintado?      ", cab.fill.fgColor.rgb, "| negrito:", cab.font.b, "| cor da letra:", cab.font.color.rgb if cab.font.color else None)
z = s.cell(row=6, column=1)
print(" zebra na linha par?     ", z.fill.fgColor.rgb)
print()
print("DOCX ------------------------------------------------")
doc = docx.Document('$SC/clientes.docx')
print(" paragrafos:", [p.text for p in doc.paragraphs if p.text][:3])
t = doc.tables[0]
print(" tabela:", len(t.rows), "linhas x", len(t.columns), "colunas")
for row in t.rows[:3]:
    print("  ", [c.text for c in row.cells])
