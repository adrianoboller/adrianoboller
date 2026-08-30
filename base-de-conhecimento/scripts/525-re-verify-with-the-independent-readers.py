# Re-verify with the independent readers
# 28/08 16:58

import openpyxl, datetime, docx
s = openpyxl.load_workbook('$SC/clientes.xlsx').active
cab = s.cell(row=4, column=1)
print("XLSX cabeçalho: fundo", cab.fill.fgColor.rgb, "| negrito", cab.font.b, "| letra", cab.font.color.rgb)
print("XLSX zebra    : fundo", s.cell(row=6, column=1).fill.fgColor.rgb)
print("XLSX normal   : fundo", s.cell(row=5, column=1).fill.fgColor.rgb)
print("XLSX data     :", s.cell(row=5,column=5).value, "|", s.cell(row=5,column=5).number_format)
print("XLSX instante :", s.cell(row=5,column=6).value, "|", s.cell(row=5,column=6).number_format)
print("XLSX filtro   :", s.auto_filter.ref, "| congelado:", s.freeze_panes)
print()
d = docx.Document('$SC/clientes.docx')
t = d.tables[0]
print("DOCX tabela:", len(t.rows), "linhas x", len(t.columns), "colunas")
for row in t.rows[:2]:
    print("  ", [c.text for c in row.cells])
print("DOCX acentuação chegou?", "acentuação" in t.rows[3].cells[1].text)
