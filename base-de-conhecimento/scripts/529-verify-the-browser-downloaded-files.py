# Verify the browser-downloaded files
# 28/08 17:00

import openpyxl, docx
s = openpyxl.load_workbook('$SC/web-xlsx.xlsx').active
print("baixado do navegador, lido pelo openpyxl:")
for r in s.iter_rows(min_row=1, max_row=8, values_only=True): print("  ", r)
print("  filtro:", s.auto_filter.ref, "| congelado:", s.freeze_panes,
      "| cabecalho:", s.cell(row=4,column=1).fill.fgColor.rgb)
d = docx.Document('$SC/web-docx.docx')
print("docx:", len(d.tables[0].rows), "linhas x", len(d.tables[0].columns), "colunas")
